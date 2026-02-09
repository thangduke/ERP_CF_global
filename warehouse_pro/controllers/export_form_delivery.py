from odoo import http
from odoo.http import request, content_disposition
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from io import BytesIO
import datetime

# Vietnamese number to words converter
def number_to_words_vi(n):
    if n == 0:
        return "không đồng"

    units = ["", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]
    teens = ["mười", "mười một", "mười hai", "mười ba", "mười bốn", "mười lăm", "mười sáu", "mười bảy", "mười tám", "mười chín"]
    tens = ["", "mười", "hai mươi", "ba mươi", "bốn mươi", "năm mươi", "sáu mươi", "bảy mươi", "tám mươi", "chín mươi"]
    hundreds = ["không", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]
    thousands = ["", "nghìn", "triệu", "tỷ"]

    def read_group_of_three(group):
        h, t, u = group // 100, (group % 100) // 10, group % 10
        result = []
        if h > 0:
            result.extend([hundreds[h], "trăm"])
        if t > 1:
            result.append(tens[t])
            if u == 1:
                result.append("mốt")
            elif u > 1:
                result.append(units[u])
        elif t == 1:
            result.append(teens[u])
        elif u > 0:
            if h > 0 or len(result) > 0: # Need to check if there are preceding hundreds
                result.append("linh")
            result.append(units[u])
        return " ".join(result)

    if n < 0:
        return "âm " + number_to_words_vi(abs(n))

    parts = []
    num_str = str(int(n))
    while len(num_str) % 3 != 0 and len(num_str) > 3:
        num_str = '0' + num_str

    groups = [int(num_str[i:i+3]) for i in range(0, len(num_str), 3)]
    
    if len(groups) == 1:
        parts.append(read_group_of_three(groups[0]))
    else:
        for i, group in enumerate(groups):
            if group > 0:
                parts.append(read_group_of_three(group))
                parts.append(thousands[len(groups) - 1 - i])

    # Clean up "không trăm linh" cases and join
    full_str = " ".join(filter(None, parts)).strip()
    # Capitalize first letter
    return (full_str[0].upper() + full_str[1:] + " đồng").replace("  ", " ")


class MaterialDeliveryExport(http.Controller):
    
    @http.route('/export/form_delivery/<int:delivery_id>', type='http', auth='user')
    def export_form_delivery(self, delivery_id, **kwargs):
        delivery = request.env['material.delivery'].browse(delivery_id)
        if not delivery.exists():
            return request.not_found()

        company = request.env.company
        wb = Workbook()
        ws = wb.active
        ws.title = "Phiếu Xuất Kho"

        # ======= Styles =======
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right = Alignment(horizontal='right', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        # ======= Column Widths =======
        col_widths = {'A': 5, 'B': 25, 'C': 15, 'D': 12, 'E': 8, 'F': 10, 'G': 12, 'H': 18, 'I': 18, 'J': 10, 'K': 12, 'L': 15}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
            
        # ======= Content =======
        row = 1
        max_col = 12

        # --- Header ---
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        company_name = delivery.store_id.company if delivery.store_id and delivery.store_id.company else ''
        ws.cell(row, 1, company_name.upper()).font = Font(bold=True, size=12)
        ws.cell(row, 1).alignment = center
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        company_address = delivery.store_id.address if delivery.store_id and delivery.store_id.address else ''
        ws.cell(row, 1, company_address).alignment = center
        row += 2

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        ws.cell(row, 1, "PHIẾU XUẤT KHO").font = Font(bold=True, size=14)
        ws.cell(row, 1).alignment = center
        row += 1

        export_date = delivery.date_delivery or datetime.date.today()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        ws.cell(row, 1, f"Ngày {export_date.day:02d} tháng {export_date.month:02d} năm {export_date.year}").alignment = center
        row += 2

        # --- General Info ---
        ws.cell(row, 1, "Người nhận hàng:")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max_col)
        ws.cell(row, 2, delivery.receiver_id or '')
        row += 1

        ws.cell(row, 1, "Bộ phận:")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row, 2, delivery.production_name or '')
        ws.cell(row, 6, "Số phiếu:")
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
        ws.cell(row, 7, delivery.delivery_no or '')
        ws.cell(row, 10, "Ngày xuất:")
        ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=12)
        ws.cell(row, 11, delivery.date_delivery.strftime("%d/%m/%Y") if delivery.date_delivery else '')
        row += 1

        ws.cell(row, 1, "Nội dung:")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max_col)
        ws.cell(row, 2, delivery.purpose or '')
        row += 1

        ws.cell(row, 1, "Kho xuất:")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row, 2, delivery.store_id.name if delivery.store_id else '')
        ws.cell(row, 6, "Đơn hàng:")
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=max_col)
        ws.cell(row, 7, delivery.order_id.name if hasattr(delivery, 'order_id') and delivery.order_id else '')
        row += 2

        # --- Style Lines Table ---
        if delivery.style_line_ids:
            style_header = ["Mã hàng", "Số lượng sản phẩm"]
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell1 = ws.cell(row, 1, style_header[0]); cell1.font = bold; cell1.alignment = center; cell1.border = thin_border; cell1.fill = light_blue_fill
            for i in range(2, 7): ws.cell(row, i).border = thin_border; ws.cell(row, i).fill = light_blue_fill
            
            ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=12)
            cell2 = ws.cell(row, 7, style_header[1]); cell2.font = bold; cell2.alignment = center; cell2.border = thin_border; cell2.fill = light_blue_fill
            for i in range(8, 13): ws.cell(row, i).border = thin_border; ws.cell(row, i).fill = light_blue_fill
            row += 1

            for line in delivery.style_line_ids:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                ws.cell(row, 1, line.product_code_id.name if line.product_code_id else '').alignment = left
                
                ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=12)
                ws.cell(row, 7, line.quantity).alignment = right
                
                for col in range(1, 13): ws.cell(row, col).border = thin_border
                row += 1
            row += 1


        # --- Details Table Header ---
        header = ["STT", "Mtr#", "Mtr code", "Mtr Type", "Unit","Dimension","Color#","Color Name", "Suppier", "SL xuất", "Price$", "Thành tiền"]
        for i, header_title in enumerate(header):
            col = i + 1
            cell = ws.cell(row, col, header_title)
            cell.font = bold
            cell.alignment = center
            cell.border = thin_border
            cell.fill = light_blue_fill
        row += 1

        # --- Details Table Body ---
        total = 0
        for idx, line in enumerate(delivery.delivery_line_ids, 1):
            ws.cell(row, 1, idx)
            ws.cell(row, 2, line.mtr_name)
            ws.cell(row, 3, line.mtr_code)
            ws.cell(row, 4, line.mtr_type.name or '')
            ws.cell(row, 5, line.rate or '')
            ws.cell(row, 6, line.dimension or '')
            ws.cell(row, 7, line.color_item or '')
            ws.cell(row, 8, line.color_name or '')
            ws.cell(row, 9, line.supplier.name_supplier or '')
            ws.cell(row, 10, line.qty)
            ws.cell(row, 11, line.price)
            ws.cell(row, 12, line.subtotal)
            total += line.subtotal

            for col in range(1, 13):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if col == 1: 
                    cell.alignment = center
                elif col in [2, 3, 4, 5, 6, 7, 8, 9]: 
                    cell.alignment = left
                else: 
                    cell.alignment = right
                    if col > 9:
                        cell.number_format = '#,##0.00'
            row += 1
            
        # --- Footer ---
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        total_cell = ws.cell(row, 1, "TỔNG CỘNG:")
        total_cell.font = bold
        total_cell.alignment = right
        
        val_cell = ws.cell(row, 12, total)
        val_cell.font = bold
        val_cell.alignment = right
        val_cell.number_format = '#,##0.00'
        
        for col in range(1, 13): ws.cell(row, col).border = thin_border
        row += 1

        ws.merge_cells(f'A{row}:L{row}')
        ws[f'A{row}'] = f"Tổng cộng số tiền (Viết bằng chữ): {number_to_words_vi(total)}"
        ws[f'A{row}'].font = Font(italic=True)
        row += 1

        ws.merge_cells(f'A{row}:L{row}')
        ws[f'A{row}'] = "Số chứng từ gốc kèm theo:"
        row += 2

        ws.merge_cells(f'J{row}:L{row}')
        ws[f'J{row}'].value = f"Ngày {export_date.day:02d} tháng {export_date.month:02d} năm {export_date.year}"
        ws[f'J{row}'].alignment = center
        row += 1
        
        # --- Signatures ---
        signature_titles = {'A': "Người lập phiếu", 'D': "Người nhận hàng", 'G': "Thủ kho", 'J': "Giám đốc"}
        for col_start_letter, title in signature_titles.items():
            merge_range = f'{col_start_letter}{row}:{chr(ord(col_start_letter)+2)}{row}'
            ws.merge_cells(merge_range)
            cell = ws[f'{col_start_letter}{row}']; cell.value = title; cell.alignment = center; cell.font = bold

            subtitle_merge_range = f'{col_start_letter}{row + 1}:{chr(ord(col_start_letter)+2)}{row + 1}'
            ws.merge_cells(subtitle_merge_range)
            subtitle_cell = ws[f'{col_start_letter}{row + 1}']; subtitle_cell.value = "(ký, họ tên)"; subtitle_cell.alignment = center

        # Add names a few rows below for signature
        name_row = row + 4
        signature_names = {
            'A': delivery.employee_id.name if delivery.employee_id else '',
            'D': delivery.receiver_id or '',
            'G': delivery.storekeeper_id.name if delivery.storekeeper_id else '',
            'J': delivery.director_id.name if delivery.director_id else ''
        }
        for col_start_letter, name in signature_names.items():
            name_merge_range = f'{col_start_letter}{name_row}:{chr(ord(col_start_letter)+2)}{name_row}'
            ws.merge_cells(name_merge_range)
            name_cell = ws[f'{col_start_letter}{name_row}']
            name_cell.value = name
            name_cell.alignment = center
            name_cell.font = bold
            
            
        # ======= Export File =======
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"PhieuXuatKho_{delivery.delivery_no or 'NoNumber'}.xlsx"

        return request.make_response(
            stream.getvalue(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition(filename))
            ]
        )