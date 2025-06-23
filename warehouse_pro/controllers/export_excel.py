from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl import load_workbook
import os

class ShelfExportController(http.Controller):

    @http.route('/export/material/<int:shelf_id>', type='http', auth='user')
    def export_shelf_material_excel(self, shelf_id, **kwargs):
        shelf = request.env['shelf.list'].sudo().browse(shelf_id)
        lines = shelf.shelf_line_ids

        # Xây dựng đường dẫn tuyệt đối đến file mẫu
        module_path = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(module_path, '../static/src/xlsx/export_shelf.xlsx')

        # Kiểm tra nếu file không tồn tại
        if not os.path.exists(template_path):
            return request.not_found()

        # Load file mẫu
        workbook = load_workbook(template_path)
        sheet = workbook.active  # Hoặc workbook['TênSheet'] nếu bạn dùng nhiều sheet

        start_row = 21

        # Chèn dòng mới bắt đầu từ dòng 21, đủ cho số dòng cần ghi
        sheet.insert_rows(start_row, amount=len(lines))

        # Ghi dữ liệu từ dòng 21 trở đi
        for idx, line in enumerate(lines):
            row = start_row + idx
            sheet.cell(row=row, column=1, value=line.position or '')
            sheet.cell(row=row, column=2, value=line.mtr_type.name or '')
            sheet.cell(row=row, column=3, value=line.mtr_no or '')
            sheet.cell(row=row, column=4, value=line.mtr_code or '')
            sheet.cell(row=row, column=5, value=line.mtr_name or '')
            sheet.cell(row=row, column=6, value=line.dimension or '')
            sheet.cell(row=row, column=7, value=line.color_item or '')
            sheet.cell(row=row, column=8, value=line.color_name or '')
            sheet.cell(row=row, column=9, value=line.est_qty or 0)
            sheet.cell(row=row, column=10, value=line.act_qty or 0)
            sheet.cell(row=row, column=11, value=line.rate or '')
            sheet.cell(row=row, column=12, value=line.price or 0)
            sheet.cell(row=row, column=13, value=line.supplier or '')
            sheet.cell(row=row, column=14, value=line.country or '')
            '''
            sheet.cell(row=row, column=15, value=line.cif_price or 0)
            sheet.cell(row=row, column=16, value=line.fob_price or 0)
            sheet.cell(row=row, column=17, value=line.exwork_price or 0)
            sheet.cell(row=row, column=18, value=line.total or 0)
            '''
        # Lưu vào bộ nhớ và trả về file
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        current_time = datetime.now().strftime('%H.%M.%S.%d.%m')
        filename = f'vat_tu_trong_ke_{current_time}.xlsx'

        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename={filename}')
            ]
        )