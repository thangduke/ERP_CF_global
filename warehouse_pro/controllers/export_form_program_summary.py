from odoo import http
from odoo.http import request, content_disposition
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import json
from itertools import groupby

class StockProgramSummaryExport(http.Controller):
    @http.route('/export/stock_program_summary', type='http', auth='user')
    def export_stock_program_summary(self, ids=None, domain=None, **kwargs):
        if ids:
            ids = json.loads(ids)
            records = request.env['material.stock.program.summary'].browse(ids)
        elif domain:
            domain = json.loads(domain)
            records = request.env['material.stock.program.summary'].search(domain)
        else:
            records = request.env['material.stock.program.summary'].search([])

        records = records.sorted(key=lambda r: r.order_id.name or '')
        groups = [(k, list(g)) for k, g in groupby(records, key=lambda r: r.order_id.name or 'Chưa phân loại')]

        workbook = Workbook()

        header_font = Font(bold=True, name='Times New Roman', size=12)
        cell_font = Font(name='Times New Roman', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        headers = [
            "STT", "Mtr#", "Mtr name", "Unit", "Mtr Type", "Mtr code", "Dimension", "Color#", "Color name", "Supplier","Price $",
            "SL tồn đầu", "Dư đầu $", "SL nhập", "Tiền nhập $",
            "SL xuất", "Tiền xuất $", "SL tồn cuối", "Dư cuối $"
        ]
        
        num_cols = len(headers)

        if len(groups) > 1:
            # More than one program, create a sheet for each
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook.active)

            for program_name, group in groups:
                sheet = workbook.create_sheet(title=program_name[:31])
                
                sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
                sheet['A1'] = 'BÁO CÁO TỒN KHO THEO CHƯƠNG TRÌNH'
                sheet['A1'].font = Font(bold=True, name='Times New Roman', size=16)
                sheet['A1'].alignment = alignment_center
                sheet.row_dimensions[1].height = 30

                sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
                today = datetime.date.today()
                sheet['A2'] = f"Ngày {today.day} tháng {today.month} năm {today.year}"
                sheet['A2'].alignment = alignment_center
                sheet['A2'].font = cell_font

                row_index = 4

                sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=num_cols)
                program_cell = sheet.cell(row=row_index, column=1, value=f"Chương trình: {program_name}")
                program_cell.font = header_font
                program_cell.alignment = alignment_left
                row_index += 1

                header_row_idx = row_index
                for c_idx, header_title in enumerate(headers, 1):
                    cell = sheet.cell(row=header_row_idx, column=c_idx, value=header_title)
                    cell.font = header_font
                    cell.alignment = alignment_center
                    cell.border = border
                    cell.fill = header_fill
                row_index += 1

                group_start_row = row_index
                stt = 1
                for record in group:
                    col = 1
                    sheet.cell(row=row_index, column=col, value=stt).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.name).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.mtr_name).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.rate).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.mtr_type.name if record.mtr_type else '').border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.mtr_code).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.dimension).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.color_item).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.color_name).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.supplier.name_supplier if record.supplier else '').border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.price).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.qty_opening).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.value_opening).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.qty_in).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.value_in).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.qty_out).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.value_out).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.qty_closing).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.value_closing).border = border; col+=1
                    stt += 1
                    row_index += 1

                total_row = row_index
                sheet.cell(row=total_row, column=9, value="Tổng cộng").font = header_font
                sheet.cell(row=total_row, column=9).border = border
                for col in range(10, num_cols + 1):
                    col_letter = get_column_letter(col)
                    cell = sheet.cell(row=total_row, column=col, value=f"=SUM({col_letter}{group_start_row}:{col_letter}{row_index - 1})")
                    cell.font = header_font
                    cell.border = border
                
                for col_idx in range(1, num_cols + 1):
                    sheet.column_dimensions[get_column_letter(col_idx)].auto_size = True
        else:
            # Original logic for one or zero programs
            sheet = workbook.active
            sheet.title = "Báo cáo tồn kho theo chương trình"
            
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            sheet['A1'] = 'BÁO CÁO TỒN KHO THEO CHƯƠNG TRÌNH'
            sheet['A1'].font = Font(bold=True, name='Times New Roman', size=16)
            sheet['A1'].alignment = alignment_center
            sheet.row_dimensions[1].height = 30

            sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
            today = datetime.date.today()
            sheet['A2'] = f"Ngày {today.day} tháng {today.month} năm {today.year}"
            sheet['A2'].alignment = alignment_center
            sheet['A2'].font = cell_font

            row_index = 4

            for program_name, group in groups:
                sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=num_cols)
                program_cell = sheet.cell(row=row_index, column=1, value=f"Chương trình: {program_name}")
                program_cell.font = header_font
                program_cell.alignment = alignment_left
                row_index += 1

                header_row_idx = row_index
                for c_idx, header_title in enumerate(headers, 1):
                    cell = sheet.cell(row=header_row_idx, column=c_idx, value=header_title)
                    cell.font = header_font
                    cell.alignment = alignment_center
                    cell.border = border
                    cell.fill = header_fill
                row_index += 1

                group_start_row = row_index
                stt = 1
                for record in group:
                    col = 1
                    sheet.cell(row=row_index, column=col, value=stt).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.mtr_no).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.mtr_name).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.rate).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.mtr_type.name if record.mtr_type else '').border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.mtr_code).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.dimension).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.color_item).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.color_name).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.supplier.name_supplier if record.supplier else '').border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.price).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.qty_opening).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.value_opening).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.qty_in).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.value_in).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.qty_out).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.value_out).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.qty_closing).border = border; col+=1
                    sheet.cell(row=row_index, column=col, value=record.value_closing).border = border; col+=1
                    
                    stt += 1
                    row_index += 1

                total_row = row_index
                sheet.cell(row=total_row, column=9, value="Tổng cộng").font = header_font
                sheet.cell(row=total_row, column=9).border = border
                for col in range(10, num_cols + 1):
                    col_letter = get_column_letter(col)
                    cell = sheet.cell(row=total_row, column=col, value=f"=SUM({col_letter}{group_start_row}:{col_letter}{row_index - 1})")
                    cell.font = header_font
                    cell.border = border
                
                row_index += 2
            
            for col_idx in range(1, num_cols + 1):
                sheet.column_dimensions[get_column_letter(col_idx)].auto_size = True

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                for cell in row:
                    if cell.font != header_font and cell.font.sz != 16:
                         cell.font = cell_font

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        file_name = f"Baocao_ton_kho_chuong_trinh_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        return request.make_response(
            buffer.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition(file_name))
            ]
        )