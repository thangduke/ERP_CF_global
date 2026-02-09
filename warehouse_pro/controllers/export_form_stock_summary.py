from odoo import http
from odoo.http import request, content_disposition
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from io import BytesIO
import datetime
import json

class StockSummaryExport(http.Controller):
    
    @http.route('/export/stock_summary', type='http', auth='user')
    def export_stock_summary(self, model, ids=None, domain=None, **kwargs):
        
        if ids:
            ids = [int(i) for i in ids.split(',')]
            records = request.env[model].browse(ids)
        elif domain:
            records = request.env[model].search(json.loads(domain))
        else:
            records = request.env[model].search([])

        wb = Workbook()
        ws = wb.active
        ws.title = "Báo cáo NXT"

        # Styles
        title_font = Font(bold=True, size=16, name='Times New Roman')
        date_font = Font(name='Times New Roman', size=11)
        header_font = Font(bold=True, name='Times New Roman', size=12)
        cell_font = Font(name='Times New Roman', size=11)
        bold_cell_font = Font(bold=True, name='Times New Roman', size=11)

        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        # Report Header
        ws.merge_cells('A1:S1')
        ws['A1'] = "BÁO CÁO TỔNG HỢP TỒN KHO "
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 30
        
        ws.merge_cells('A2:S2')
        today = datetime.date.today()
        ws['A2'] = f"Ngày {today.day} tháng {today.month} năm {today.year}"
        ws['A2'].alignment = center_align
        ws['A2'].font = date_font

        # Column Headers
        headers = [
            ("STT", 1), ("Mtr#", 1), ("Mtr name", 1), ("Unit", 1),
            ("Mtr Type", 1), ("Mtr code", 1), ("Dimension", 1), ("Color#", 1), ("Color name", 1), 
            ("Supplier", 1), ("Price $", 1),
            ("Tồn đầu kỳ", 2), ("Nhập trong kỳ", 2), ("Xuất trong kỳ", 2), ("Tồn cuối kỳ", 2)
        ]
        
        
        col_idx = 1
        for title, span in headers:
            if span == 1:
                ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
                ws.cell(row=3, column=col_idx, value=title)
            else:
                ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx + span - 1)
                ws.cell(row=3, column=col_idx, value=title)
                ws.cell(row=4, column=col_idx, value="SL")
                ws.cell(row=4, column=col_idx + 1, value="Giá trị")
            col_idx += span

        for row_cells in ws['A3:S4']:
            for cell in row_cells:
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = header_fill

        # Data rows
        row_num = 5
        totals = {
            'qty_opening': 0, 'value_opening': 0, 'qty_in': 0, 'value_in': 0,
            'qty_out': 0, 'value_out': 0, 'qty_closing': 0, 'value_closing': 0,
        }
        for idx, rec in enumerate(records, 1):
            data_row = [
                idx,
                rec.name,
                rec.mtr_name,
                rec.rate,
                rec.mtr_type.name if rec.mtr_type else '',
                rec.mtr_code,
                rec.dimension,
                rec.color_item,
                rec.color_name,
                rec.supplier.name_supplier if rec.supplier else '',
                rec.price,
                rec.qty_opening,
                rec.value_opening,
                rec.qty_in,
                rec.value_in,
                rec.qty_out,
                rec.value_out,
                rec.qty_closing,
                rec.value_closing,
            ]
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                cell.font = cell_font
                
                # Alignment and formatting based on column content
                if col in [1, 4]:  # STT, Unit
                    cell.alignment = center_align
                elif col in [2, 3, 5, 6, 7, 8, 9, 10]:  # Text fields
                    cell.alignment = left_align
                else:  # Numeric fields (Price and stock values)
                    cell.number_format = '#,##0.00'
                    cell.alignment = right_align

            for key in totals:
                totals[key] += getattr(rec, key, 0)
            row_num += 1

        # Total Row
        ws.merge_cells(f'A{row_num}:K{row_num}')
        ws[f'A{row_num}'] = "Tổng cộng"
        ws[f'A{row_num}'].font = bold_cell_font
        ws[f'A{row_num}'].alignment = center_align
        
        total_data = [
            totals['qty_opening'], totals['value_opening'],
            totals['qty_in'], totals['value_in'],
            totals['qty_out'], totals['value_out'],
            totals['qty_closing'], totals['value_closing'],
        ]
        for col, value in enumerate(total_data, 12):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.font = bold_cell_font
            cell.number_format = '#,##0.00'
            cell.alignment = right_align
        
        for i in range(1, 20):
            ws.cell(row=row_num, column=i).border = thin_border

        # Column widths
        widths = {
            'A': 5, 'B': 15, 'C': 40, 'D': 8, 'E': 15, 'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 20, 'K': 15,
            'L': 15, 'M': 18, 'N': 15, 'O': 18, 'P': 15, 'Q': 18, 'R': 15, 'S': 18
        }
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

        # Return file
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"BaoCao_TON_KHO_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        return request.make_response(
            stream.getvalue(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition(filename))
            ]
        )