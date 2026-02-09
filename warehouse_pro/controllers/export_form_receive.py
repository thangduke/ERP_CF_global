from odoo import http
from odoo.http import request, content_disposition
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from io import BytesIO
import datetime

class MaterialReceiveExport(http.Controller):

    @http.route('/export/form_receive/<int:receive_id>', type='http', auth='user')
    def export_form_receive(self, receive_id, **kwargs):
        receive = request.env['material.receive'].browse(receive_id)
        if not receive.exists():
            return request.not_found()

        wb = Workbook()
        ws = wb.active
        ws.title = "Phiếu nhập kho"

        # Font và style
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        right = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ========================= HEADER =========================
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{receive.store_id.company}"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = center

        ws.merge_cells('A2:H2')
        ws['A2'] = f"{receive.store_id.address}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = center

        ws.merge_cells('A4:H4')
        ws['A4'] = "PHIẾU NHẬP KHO"
        ws['A4'].font = Font(bold=True, size=14)
        ws['A4'].alignment = center

        ws.merge_cells('A5:H5')
        ws['A5'] = f"Ngày {datetime.date.today().day:02d} tháng {datetime.date.today().month:02d} năm {datetime.date.today().year}"
        ws['A5'].alignment = center

        # ========================= THÔNG TIN =========================
        ws['A7'] = "Đơn vị:"
        ws['B7'] = receive.po_id.name or ''
        ws['A8'] = "Người giao hàng:"
        ws['B8'] = receive.deliver_id.name or ''
        ws['D8'] = "Số:"
        ws['E8'] = receive.receipt_no
        ws['G8'] = "Ngày:"
        ws['H8'] = receive.date_create.strftime("%d/%m/%Y") if receive.date_create else ''
        ws['A9'] = "Nội dung:"
        ws['B9'] = receive.purpose or ''
        ws['A10'] = "Đơn hàng:"
        ws['B10'] = receive.order_id.name or ''
        ws['D10'] = "Kho hàng:"
        ws['E10'] = receive.store_id.name or ''

        # ========================= BẢNG CHI TIẾT =========================
        header = ["STT", "Tên vật tư, quy cách", "Mã vật tư", "ĐVT", "Theo C.Từ", "Thực nhập", "Đơn giá", "Thành tiền"]
        light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        
        for col, title in enumerate(header, 1):
            cell = ws.cell(row=11, column=col, value=title)
            cell.font = bold
            cell.alignment = center
            cell.border = thin_border
            cell.fill = light_blue_fill

        total = 0
        row = 12
        for idx, line in enumerate(receive.invoice_id.invoice_line_ids, 1):
            ws.cell(row=row, column=1, value=idx).alignment = center
            ws.cell(row=row, column=2, value=line.mtr_name)
            ws.cell(row=row, column=3, value=line.mtr_code)
          #   ws.cell(row=row, column=4, value=line.uom_id.name if hasattr(line, 'uom_id') else '')
          #  ws.cell(row=row, column=5, value=line.quantity or '')
            ws.cell(row=row, column=6, value=line.act_qty)
            ws.cell(row=row, column=7, value=line.price)
            amount = line.price * line.act_qty
            ws.cell(row=row, column=8, value=amount)
            total += amount
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right' if col in [6, 7, 8] else 'left')
            row += 1

        # ========================= TỔNG CỘNG =========================
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "CỘNG:"
        ws[f'A{row}'].alignment = right
        ws[f'A{row}'].font = bold
        ws[f'H{row}'] = total
        ws[f'H{row}'].font = bold
        ws[f'H{row}'].alignment = right
        row += 1

        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "Tổng cộng số tiền (Viết bằng chữ):"
        row += 1

        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = "Số chứng từ gốc kèm theo:"
        row += 2

        ws.merge_cells(f'F{row}:H{row}')
        ws[f'F{row}'].value = "Ngày .... tháng .... năm ...."
        ws[f'F{row}'].alignment = center
        row += 1
        
        # ========================= CHỮ KÝ =========================
        signature_titles = {
            'A': "Người lập phiếu",
            'C': "Người giao hàng",
            'E': "Thủ kho",
            'G': "Kế toán trưởng"
        }
        
        for col_start_letter, title in signature_titles.items():
            start_col_idx = ord(col_start_letter) - ord('A') + 1
            merge_range = f'{col_start_letter}{row}:{chr(ord(col_start_letter)+1)}{row}'
            ws.merge_cells(merge_range)
            cell = ws[f'{col_start_letter}{row}']
            cell.value = title
            cell.alignment = center
            cell.font = bold

            # Add (ký, họ tên)
            subtitle_merge_range = f'{col_start_letter}{row + 1}:{chr(ord(col_start_letter)+1)}{row + 1}'
            ws.merge_cells(subtitle_merge_range)
            subtitle_cell = ws[f'{col_start_letter}{row + 1}']
            subtitle_cell.value = "(ký, họ tên)"
            subtitle_cell.alignment = center
            
        # Căn chỉnh cột
        col_widths = [6, 45, 12, 8, 12, 12, 15, 18]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # ========================= TRẢ FILE =========================
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"PhieuNhapKho_{receive.receipt_no}.xlsx"
        return request.make_response(
            stream.getvalue(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition(filename))
            ]
        )
