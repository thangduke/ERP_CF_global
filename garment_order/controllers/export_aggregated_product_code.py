from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment
import os
from openpyxl.drawing.image import Image

class AggregatedProductColorSizeExportController(http.Controller):
    
    @http.route('/export/aggregated_product_code/<int:product_code_id>', type='http', auth='user')
    def export_aggregated_product_color_size_excel(self, product_code_id, **kwargs):
        try:
            product_code = request.env['product.code'].sudo().browse(product_code_id)
            if not product_code or not product_code.exists():
                return request.make_response(
                    "Không tìm thấy Style hoặc Style không tồn tại.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            # Đường dẫn đến file template gốc
            module_path = os.path.dirname(os.path.abspath(__file__))
            template_path = os.path.join(module_path, '../static/src/xlsx/export_material_product/export_aggregated_product_code.xlsx')
            if not os.path.exists(template_path):
                return request.make_response(
                    "Không tìm thấy file mẫu export_material_product_code.xlsx.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            wb = load_workbook(template_path)
            ws = wb.active

            # --- Ghi thông tin Style và áp dụng font ---
            font8 = Font(size=8)
            center_alignment = Alignment(horizontal='center', vertical='center')
            header_data = {
                'C7': product_code.warehouse_order_id.name or '',
                'C8': product_code.name or '',
                'C9': product_code.description or '',
                'F7': product_code.customer_id.name_customer or '',
                'F8': product_code.ean_no or '',
                'H7': product_code.total_order_qty or '',
                'H8': product_code.total_test_qty or '',
            }
            for cell, value in header_data.items():
                ws[cell] = value
                ws[cell].font = font8
                ws[cell].alignment = center_alignment

            # --- Tự động tạo header ---
            ordered_headers = [
                'Mtr#', 'Type', 'Mtr.Code', 'Material name', 'Dimension', 'Color#',
                'Color name', 'Color set', 'Rate', 'Supplier#', 'Supplier', 'Consumption', 
                'Price', 'Cif_price', 'Fob_price', 'Exwork_price', 'Total'
            ]
            
            header_font = Font(bold=True, size=8)
            header_map = {}
            start_row = 11
            
            # Xóa header cũ và ghi header mới
            for cell in ws[start_row]:
                cell.value = None
            for col_idx, header_title in enumerate(ordered_headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=header_title)
                cell.font = header_font
                header_map[header_title] = col_idx

            # Dòng bắt đầu ghi dữ liệu
            current_row = start_row + 1
            data_font = Font(size=8)
            
            # Gọi action để tính toán vật tư tổng hợp
            product_code.action_compute_grouped_materials()

            if product_code.aggregated_material_ids:
                for material in product_code.aggregated_material_ids:
                    row_data = {
                        'Mtr#': material.name or '',
                        'Type': material.mtr_type.name if material.mtr_type else '',
                        'Mtr.Code': material.mtr_code or '',
                        'Material name': material.mtr_name or '',
                        'Dimension': material.dimension or '',
                        'Color#': material.color_item or '',
                        'Color name': material.color_name or '',
                        'Color set': material.color_set or '',
                        'Rate': material.rate or '',
                        'Supplier#': material.supplier.supplier_index if material.supplier else '',
                        'Supplier': material.supplier.name_supplier if material.supplier else '',
                        'Consumption': material.cons_qty or 0,
                        'Price': material.price or 0,
                        'Cif_price': material.cif_price or 0,
                        'Fob_price': material.fob_price or 0,
                        'Exwork_price': material.exwork_price or 0,
                        'Total': material.total or 0,
                    }

                    # Ghi dòng vào excel
                    for header, col_idx in header_map.items():
                        cell = ws.cell(row=current_row, column=col_idx, value=row_data.get(header, ''))
                        cell.font = data_font
                    
                    current_row += 1


            # Lưu file vào bộ nhớ
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_data = output.read()
            wb.close()

            filename = f'Product_{product_code.name or product_code.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
            return request.make_response(file_data, headers)
        except Exception as e:
            return request.make_response(
                f"Lỗi khi xuất file Excel: {str(e)}",
                [('Content-Type', 'text/plain; charset=utf-8')]
            )