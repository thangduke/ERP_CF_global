from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import os
from openpyxl.drawing.image import Image

class OrderManagementExportController(http.Controller):
    
    @http.route('/export/order_management/<int:order_id>', type='http', auth='user')
    def export_order_management_excel(self, order_id, **kwargs):
        try:
            order = request.env['warehouse.order'].sudo().browse(order_id)
            if not order or not order.exists():
                return request.make_response(
                    "Không tìm thấy chương trình hoặc chương trình không tồn tại.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            # Đường dẫn đến file template gốc
            module_path = os.path.dirname(os.path.abspath(__file__))
            template_path = os.path.join(module_path, '../static/src/xlsx/export_order_management/export_order_management.xlsx')
            if not os.path.exists(template_path):
                return request.make_response(
                    "Không tìm thấy file mẫu export_order_management.xlsx.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            wb = load_workbook(template_path)
            ws = wb.active

            # --- Ghi thông tin chương trình vào các ô ---
            ws['C7'] = order.order_index or ''
            ws['C8'] = order.name or ''
            ws['C9'] = order.customer_id.name_customer or ''
            ws['C10'] = order.order_date.strftime('%d/%m/%Y') if order.order_date else ''
            ws['C11'] = order.ship_date.strftime('%d/%m/%Y') if order.ship_date else ''
            ws['E9'] = order.customer_id.description or ''
            ws['G9'] = order.customer_id.address or ''
            
            # Dòng bắt đầu ghi dữ liệu (ví dụ: dòng 12, sau header)
            start_row = 13
            data_font = Font(size=8)

            # Lấy danh sách Style (product.code) theo chương trình
            if not order.product_code_ids:
                return request.make_response(
                    "Chương trình chưa có Style nào.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            num_lines = sum(len(request.env['product.color.size'].sudo().search([('product_code_id', '=', code.id)])) or 1 for code in order.product_code_ids)
            ws.insert_rows(start_row, amount=num_lines)

            idx = start_row
            for code in order.product_code_ids:
                color_size_lines = request.env['product.color.size'].sudo().search([
                    ('product_code_id', '=', code.id)
                ])
                if color_size_lines:
                    for cs in color_size_lines:
                        ws[f'B{idx}'] = code.name or ''
                        ws[f'C{idx}'] = code.description or ''
                        ws[f'D{idx}'] = cs.color_id.color_code or ''
                        ws[f'E{idx}'] = cs.color_id.name or ''
                        ws[f'F{idx}'] = cs.label or ''
                        ws[f'G{idx}'] = cs.dimpk or 0
                        ws[f'H{idx}'] = cs.ppk or 0
                        ws[f'I{idx}'] = cs.size_id.name or ''
                        ws[f'J{idx}'] = cs.order_qty or 0
                        ws[f'K{idx}'] = cs.test_qty or 0
                        ws[f'L{idx}'] = cs.unit_cost or 0
                        ws[f'M{idx}'] = cs.ext or 0
                        # Áp dụng font cho các ô trong dòng
                        for col_num in range(2, 14):  # Từ cột B đến M
                            ws.cell(row=idx, column=col_num).font = data_font

                        idx += 1
                else:
                    ws[f'B{idx}'] = code.ean_no or ''
                    ws[f'C{idx}'] = code.description or ''
                    
                    # Áp dụng font cho các ô trong dòng
                    ws.cell(row=idx, column=2).font = data_font
                    ws.cell(row=idx, column=3).font = data_font
                    # Các cột còn lại để trống
                    idx += 1
                    
            desc_row = idx + 1  # hoặc idx nếu muốn ngay sau dữ liệu
            # Tiêu đề "Ghi chú:"
            ws[f'B{desc_row}'] = "Ghi chú:"
            ws[f'B{desc_row}'].font = Font(bold=True, italic=True, color="FF0000")  # Đỏ, đậm, nghiêng

            # Ghi mô tả ở cột C (hoặc M nếu bạn muốn)
            ws.merge_cells(f'C{desc_row}:M{desc_row}')
            ws[f'C{desc_row}'] = order.description or ''
            ws[f'C{desc_row}'].font = Font(italic=True)
            # Đóng khung cho vùng mô tả
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col in range(3, 14):  # C đến M là 3 đến 13
                ws[f'{get_column_letter(col)}{desc_row}'].border = border

            ws[f'C{desc_row}'].alignment = ws[f'C{desc_row}'].alignment.copy(horizontal='left', vertical='center')
            
            # Lưu file vào bộ nhớ
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_data = output.read()
            wb.close()

            # Trả file về cho người dùng
            filename = f'Product_{order.name or order.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

            headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
            return request.make_response(file_data, headers)
        except Exception as e:
            return request.make_response(
                f"Lỗi khi xuất file Excel: {str(e)}",
                [('Content-Type', 'text/plain; charset=utf-8')]
            )