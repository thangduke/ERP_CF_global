from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import os

class ExportPriceCalculationController(http.Controller):

    @http.route('/export/price_calculation/<int:price_calculation_id>', type='http', auth='user')
    def export_price_calculation_excel(self, price_calculation_id, **kwargs):
        try:
            price_calc = request.env['product.price.calculation'].sudo().browse(price_calculation_id)
            if not price_calc or not price_calc.exists():
                return request.make_response(
                    "Không tìm thấy bảng tính giá hoặc bảng tính giá không tồn tại.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            # Tạo workbook và worksheet mới
            wb = Workbook()
            ws = wb.active
            ws.title = "Price Calculation"

            # Định nghĩa các style
            header_font = Font(bold=True, size=10)
            data_font = Font(size=10)
            bold_font = Font(bold=True)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

            # Thiết lập độ rộng cột
            column_widths = {'A': 15, 'B': 15, 'C': 25, 'D': 40, 'E': 15, 'F': 15, 'G': 25, 'H': 10, 'I': 30, 'J': 15, 'K': 15, 'L': 15}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # --- Tiêu đề chính ---
            ws.merge_cells('A1:L1')
            title_cell = ws['A1']
            title_cell.value = "BẢNG TÍNH GIÁ SẢN PHẨM"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30

            # --- Thông tin chung ---
            general_info_labels = {
                'B4': 'Mã tính giá:', 'B5': 'Chương trình:', 'B6': 'Khách hàng:', 'B7': 'Style:', 'B8': 'Color / Size:',
                'F4': 'Ngày tính giá:', 'F5': 'Người tạo:'
            }
            for cell, label in general_info_labels.items():
                ws[cell] = label
                ws[cell].font = header_font

            ws['C4'] = price_calc.name or ''
            ws['C5'] = price_calc.warehouse_order_id.name or ''
            ws['C6'] = price_calc.customer_id.name_customer or ''
            ws['C7'] = price_calc.product_code_id.name or ''
            ws['C8'] = f"{price_calc.product_color_size_id.color_id.name} / {price_calc.product_color_size_id.size_id.name}"
            ws['G4'] = price_calc.date_calculation.strftime('%d/%m/%Y') if price_calc.date_calculation else ''
            ws['G5'] = price_calc.employee_id.name or ''

            # --- Tiêu đề chi tiết vật tư ---
            material_header = ['Position', 'Mtr#', 'Mtr Type', 'Mtr Name', 'Mtr Code', 'Dimension', 'Color Name', 'Rate', 'Supplier', 'Consumption', 'Price', 'Total Price']
            start_row_header = 11
            for i, header_title in enumerate(material_header, start=1):
                cell = ws.cell(row=start_row_header, column=i, value=header_title)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Ghi chi tiết vật tư
            start_row_data = 12
            idx = start_row_data
            material_cost = 0.0
            
            for line in price_calc.calculation_line_ids:
                total_price = line.total_price or 0
                material_cost += total_price
                data_row = [
                    line.position or '',
                    line.name or '',
                    line.mtr_type.name if line.mtr_type else '',
                    line.mtr_name or '',
                    line.mtr_code or '',
                    line.dimension or '',
                    line.color_name or '',
                    line.rate or '',
                    line.supplier.name_supplier if line.supplier else '',
                    line.consumption or 0,
                    line.price or 0,
                    total_price
                ]
                for col_num, value in enumerate(data_row, start=1):
                    cell = ws.cell(row=idx, column=col_num, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    if col_num == 10: # Consumption
                        cell.number_format = '#,##0.00'
                    elif col_num in [11, 12]: # Price, Total Price
                        cell.number_format = '#,##0.000'
                idx += 1
            
            if idx == start_row_data:
                idx += 1 

            # --- COST SUMMARY ---
            summary_start_row = idx + 1
            ws.merge_cells(f'B{summary_start_row}:H{summary_start_row}')
            summary_title_cell = ws[f'B{summary_start_row}']
            summary_title_cell.value = "COST SUMMARY"
            summary_title_cell.font = Font(bold=True, size=12)
            summary_title_cell.alignment = Alignment(horizontal='left')

            # Tính toán trước tất cả các giá trị summary trong Python
            waste_cost = material_cost * (price_calc.waste_percent / 100.0) if price_calc.waste_percent else 0
            finance_cost = material_cost * (price_calc.finance_percent / 100.0) if price_calc.finance_percent else 0
            total_net = material_cost + waste_cost + finance_cost
            cm_cost = price_calc.cut_make or 0
            admin_cost = (total_net + cm_cost) * (price_calc.admin_percent / 100.0) if price_calc.admin_percent else 0
            inspection_cost = price_calc.inspection_cost or 0
            testing_cost = price_calc.test_cost or 0
            import_export_cost = price_calc.import_export_cost or 0
            standard_fob = total_net + cm_cost + admin_cost + inspection_cost + testing_cost + import_export_cost
            surcharge_cost = standard_fob * (price_calc.surcharge_percent / 100.0) if price_calc.surcharge_percent else 0
            extra_cost = price_calc.extra_cost or 0
            final_fob = standard_fob + surcharge_cost + extra_cost
            agreed_fob = price_calc.agreed_fob or 0

            summary_data = [
                # Left Column
                {'label': '1. Material cost:', 'value': material_cost, 'row_offset': 2, 'label_col': 'B', 'value_col': 'D'},
                {'label': f'2. Waste (1 x {price_calc.waste_percent or 0:.2f}%):', 'value': waste_cost, 'row_offset': 3, 'label_col': 'B', 'value_col': 'D'},
                {'label': f'3. Finance (1 x {price_calc.finance_percent or 0:.2f}%):', 'value': finance_cost, 'row_offset': 4, 'label_col': 'B', 'value_col': 'D'},
                {'label': '4. Total net (1+2+3):', 'value': total_net, 'row_offset': 5, 'label_col': 'B', 'value_col': 'D'},
                {'label': '5. CM', 'value': cm_cost, 'row_offset': 6, 'label_col': 'B', 'value_col': 'D'},
                {'label': f'6. Admin (4&5 x {price_calc.admin_percent or 0:.2f}%):', 'value': admin_cost, 'row_offset': 7, 'label_col': 'B', 'value_col': 'D'},
                {'label': '7. Inspection cost:', 'value': inspection_cost, 'row_offset': 8, 'label_col': 'B', 'value_col': 'D'},
                # Right Column
                {'label': '8. Testing cost:', 'value': testing_cost, 'row_offset': 2, 'label_col': 'F', 'value_col': 'H'},
                {'label': '9. Import/Export cost:', 'value': import_export_cost, 'row_offset': 3, 'label_col': 'F', 'value_col': 'H'},
                {'label': '10. Standard FOB per pc:', 'value': standard_fob, 'row_offset': 4, 'label_col': 'F', 'value_col': 'H'},
                {'label': f'11. Surcharge ({price_calc.surcharge_percent or 0:.2f}%):', 'value': surcharge_cost, 'row_offset': 5, 'label_col': 'F', 'value_col': 'H'},
                {'label': '12. Extra cost per pc:', 'value': extra_cost, 'row_offset': 6, 'label_col': 'F', 'value_col': 'H'},
                {'label': '13. Final FOB (10+11+12):', 'value': final_fob, 'row_offset': 7, 'label_col': 'F', 'value_col': 'H'},
                {'label': '14. Agreed FOB per pc:', 'value': agreed_fob, 'row_offset': 8, 'label_col': 'F', 'value_col': 'H'},
            ]

            max_row_offset = 8
            for r_offset in range(2, max_row_offset + 2):
                row_idx = summary_start_row + r_offset
                if r_offset <= max_row_offset +1 :
                    ws.merge_cells(f'B{row_idx}:C{row_idx}')
                    ws[f'B{row_idx}'].border = Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    ws[f'C{row_idx}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
                    ws[f'D{row_idx}'].border = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                if r_offset <= max_row_offset:
                    ws.merge_cells(f'F{row_idx}:G{row_idx}')
                    ws[f'F{row_idx}'].border = Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    ws[f'G{row_idx}'].border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
                    ws[f'H{row_idx}'].border = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for item in summary_data:
                row = summary_start_row + item['row_offset']
                label_cell = ws[f"{item['label_col']}{row}"]
                value_cell = ws[f"{item['value_col']}{row}"]
                
                label_cell.value = item['label']
                value_cell.value = item['value']
                
                label_cell.alignment = Alignment(vertical='center', indent=1)
                value_cell.alignment = Alignment(horizontal='right', vertical='center')
                value_cell.number_format = '#,##0.00 "US$"'

                if 'Total net' in item['label'] or 'Standard FOB' in item['label'] or 'Final FOB' in item['label']:
                    label_cell.font = bold_font
                    value_cell.font = bold_font

            # Lưu file vào bộ nhớ
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_data = output.read()
            wb.close()

            filename = f'Price_Calculation_{price_calc.name or price_calc.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
            return request.make_response(file_data, headers)
        except Exception as e:
            return request.make_response(
                f"Lỗi khi xuất file Excel: {str(e)}",
                [('Content-Type', 'text/plain; charset=utf-8')]
            )