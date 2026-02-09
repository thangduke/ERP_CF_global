from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side , Alignment
import os
from openpyxl.drawing.image import Image

class MaterialPurchaseOrderExportController(http.Controller):
    
    @http.route('/export/purchase_order/<int:po_id>', type='http', auth='user')
    def export_purchase_order_excel(self, po_id, **kwargs):
        try:
            purchase_order = request.env['material.purchase.order'].sudo().browse(po_id)
            if not purchase_order.exists():
                return request.make_response("Không tìm thấy Mã lọc vật tư theo NCC.",
                                             [('Content-Type', 'text/plain; charset=utf-8')])

            # Đường dẫn đến file template
            module_path = os.path.dirname(os.path.abspath(__file__))
            template_path = os.path.join(module_path, '../static/src/xlsx/export_material_PO/export_material_purchase_order.xlsx')
            if not os.path.exists(template_path):
                return request.make_response(
                    "Không tìm thấy file mẫu export_material_purchase_order.xlsx.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            wb = load_workbook(template_path)
            ws = wb.active

            # Thêm mã chương trình vào A7 (hàng 7, cột 1)
            ws.merge_cells('C7:E7')
            ws['C7'] = purchase_order.order_id.name or ''
            ws['C7'].alignment = ws['C7'].alignment.copy(horizontal='center', vertical='center')

            # Ghi tên các PO từ C8 trở đi

            ws['C8'] = purchase_order.display_name or ''
            ws['C8'].alignment = ws['C8'].alignment.copy(horizontal='center', vertical='center')
            

            ws['C9'] = purchase_order.supplier_id.name_supplier or ''
            ws['C9'].alignment = ws['C9'].alignment.copy(horizontal='center', vertical='center')
            # Dòng bắt đầu ghi dữ liệu (ví dụ: dòng 11)
            
            # --- Tự động tạo header ---
            ordered_headers = [
                'Mtr#', 'Type', 'Mtr.Code', 'Material name', 'Dimension', 
                'Color name', 'Color set', 'Rate', 'Supplier', 'Cons. Qty', 
                'Est. Qty', 'Price', 'Cif_price', 'Fob_price', 'Exwork_price', 'Total'
            ]
            
            header_font = Font(bold=True, size=8)
            header_map = {}
            start_row = 11
            
            # Xóa header cũ và ghi header mới
            for cell in ws[start_row]:
                cell.value = None
            for col_idx, header_title in enumerate(ordered_headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=header_title)
                cell.font = header_font
                header_map[header_title] = col_idx

            # Dòng bắt đầu ghi dữ liệu
            current_row = start_row + 1
            data_font = Font(size=8)
            has_line = False

            for line in purchase_order.line_ids:
                has_line = True
                row_data = {
                    'Mtr#': line.name or '',
                    'Type': line.mtr_type.name if line.mtr_type else '',
                    'Mtr.Code': line.mtr_code or '',
                    'Material name': line.mtr_name or '',
                    'Dimension': line.dimension or '',
                    'Color#': line.color_item or '',
                    'Color name': line.color_name or '',
                    'Color set': line.color_set or '',
                    'Rate': line.rate or '',
                    'Supplier#': line.supplier.supplier_index or '',
                    'Supplier': line.supplier.name_supplier if line.supplier else '',
                    'Cons. Qty': line.cons_qty or 0,
                    'Est. Qty': line.est_qty or 0,
                    'Price': line.price or 0,
                    'Cif_price': line.cif_price or 0,
                    'Fob_price': line.fob_price or 0,
                    'Exwork_price': line.exwork_price or 0,
                    'Total': line.total or 0,
                }

                # Ghi dòng vào excel
                for header, col_idx in header_map.items():
                    cell = ws.cell(row=current_row, column=col_idx, value=row_data.get(header, ''))
                    cell.font = data_font
                
                current_row += 1

            if not has_line:
                return request.make_response(
                    "Không có dòng vật tư nào trong các PO của chương trình này.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            # Lưu file vào bộ nhớ
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_data = output.read()
            wb.close()

            filename = f'Supplier_Material_List_{purchase_order.supplier_id.name_supplier or purchase_order.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
            return request.make_response(file_data, headers)
        except Exception as e:
            return request.make_response(
                f"Lỗi khi xuất file Excel: {str(e)}",
                [('Content-Type', 'text/plain; charset=utf-8')]
            )

