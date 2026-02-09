from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from odoo.modules.module import get_module_path
import os

class WarehouseOrderExportController(http.Controller):
    _description = 'Export danh sách vật tư chi tiết của một chương trình (định dạng import)'

    @http.route('/export/warehouse_order/<int:order_id>', type='http', auth='user')
    def export_warehouse_order_excel(self, order_id, **kwargs):
        try:
            warehouse_order = request.env['warehouse.order'].sudo().browse(order_id)
            if not warehouse_order.exists():
                return request.make_response(
                    "Không tìm thấy Chương trình hoặc bản ghi không tồn tại.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            module_path = get_module_path('order_management')
            template_path = os.path.join(module_path, 'static/src/xlsx/export_order_management/export_material_warehouse_order.xlsx')
            
            if not os.path.exists(template_path):
                return request.make_response(
                    f"Không tìm thấy file mẫu tại đường dẫn: {template_path}",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            wb = load_workbook(template_path)
            ws = wb.active

            header_font = Font(bold=True, size=8, name='Arial')
            data_font = Font(size=8, name='Arial')

            # Ghi thông tin chung của chương trình
            ws['C4'] = warehouse_order.name or ''
            ws['C5'] = warehouse_order.customer_id.name_customer or ''
            ws['C6'] = warehouse_order.create_date.strftime('%d/%m/%Y') if warehouse_order.create_date else ''
            ws['D5'] = 'Program#'
            ws['E5'] = warehouse_order.order_index or ''
            
            info_cells = ['C4', 'C5', 'C6', 'E5']
            title_cells = ['D5']

            for cell_ref in info_cells:
                if ws[cell_ref]:
                    ws[cell_ref].font = data_font
            
            for cell_ref in title_cells:
                if ws[cell_ref]:
                    ws[cell_ref].font = header_font

            # --- Bắt đầu logic tổng hợp dữ liệu chi tiết ---

            product_codes = warehouse_order.product_code_ids
            if not product_codes:
                 return request.make_response(
                    "Chương trình này không có Style nào để xuất dữ liệu.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )
                 
            # 1. Xác định tất cả các size trong toàn bộ chương trình để tạo header thống nhất
            all_style_variants = product_codes.mapped('color_size_ids')
            all_sizes = all_style_variants.mapped('size_id').sorted(key=lambda s: s.name)


            # 2. Xây dựng danh sách header động
            ordered_headers = ['Style#', 'Color.style#', 'Color.style Name']
            ordered_headers.extend([f'Size_{s.name}' for s in all_sizes])
            ordered_headers.extend(['Mtr#','Type', 'Mtr.Code', 'Material name', 'Dimension'])
            ordered_headers.extend(['Color#','Color name', 'Color set', 'Rate','Supplier#' ,'Supplier'])
            ordered_headers.extend(['Price', 'Cif_price', 'Fob_price', 'Exwork_price'])
            ordered_headers.extend([f'Cons_{s.name}' for s in all_sizes])

            # 3. Ghi header mới vào dòng 7
            header_row = 7
            header_map = {}
            for cell in ws[header_row]: # Xóa header cũ
                cell.value = None
            for col_idx, header_title in enumerate(ordered_headers, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=header_title)
                cell.font = header_font
                header_map[header_title] = col_idx

            # 4. Thu thập và ghi dữ liệu cho từng style
            all_rows_data = []
            for product_code in product_codes:
                style_variants = product_code.color_size_ids
                style_colors = style_variants.mapped('color_id').sorted(key=lambda c: c.name)

                # Lấy tất cả các vật tư ('program.customer') liên quan đến style hiện tại
                related_prog_custs = request.env['program.customer'].sudo().search([
                    ('color_size_ids', 'in', style_variants.ids)
                ])

                for color in style_colors:
                    variants_for_color = style_variants.filtered(lambda v: v.color_id == color)
                    
                    relevant_prog_custs = related_prog_custs.filtered(
                        lambda pc: any(v in pc.color_size_ids for v in variants_for_color)
                    )

                    for prog_cust in relevant_prog_custs:
                        row_data = {
                            'Style#': product_code.name or '',
                            'Color.style#': color.color_code,
                            'Color.style Name': color.name,
                            'Mtr#': prog_cust.name or '',
                            'Type': prog_cust.mtr_type.name if prog_cust.mtr_type else '',
                            'Material name': prog_cust.mtr_name,
                            'Mtr.Code': prog_cust.mtr_code,
                            'Dimension': prog_cust.dimension,
                            'Color#': prog_cust.material_color_id.name if prog_cust.material_color_id else '',
                            'Color name': prog_cust.material_color_id.color_name if prog_cust.material_color_id else '',
                            'Color set': prog_cust.material_color_id.color_set_id.name if prog_cust.material_color_id.color_set_id else '',
                            'Rate': prog_cust.rate,
                            'Supplier#': prog_cust.supplier.supplier_index if prog_cust.supplier else '',
                            'Supplier': prog_cust.supplier.name_supplier if prog_cust.supplier else '',
                            'Price': prog_cust.price,
                            'Cif_price': prog_cust.cif_price,
                            'Fob_price': prog_cust.fob_price,
                            'Exwork_price': prog_cust.exwork_price,                        
                        }

                        # Lấy dữ liệu chi tiết theo từng size (dựa trên danh sách all_sizes)
                        for size in all_sizes:
                            variant = style_variants.filtered(lambda v: v.color_id == color and v.size_id == size)
                            
                            if variant and variant in prog_cust.color_size_ids:
                                size_name = size.name
                                row_data[f'Size_{size_name}'] = 'x'
                                
                                norm_line = prog_cust.norm_line_ids.filtered(lambda n: n.color_size_id == variant)
                                row_data[f'Cons_{size_name}'] = norm_line.consumption if norm_line else 0.0

                        all_rows_data.append(row_data)

            # 6. Tự động điều chỉnh độ rộng cột
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                # Thêm một chút khoảng trống để dễ đọc
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width

            # 7. Lưu và trả về file
            output = io.BytesIO()
            wb.save(output)
            file_data = output.getvalue()

            filename = f'Export_Material_{warehouse_order.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.xlsx'
            headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
            return request.make_response(file_data, headers)

        except Exception as e:
            return request.make_response(f"Lỗi khi xuất file Excel: {str(e)}", [('Content-Type', 'text/plain; charset=utf-8')])