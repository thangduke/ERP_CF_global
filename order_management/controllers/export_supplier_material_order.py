from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import os
from openpyxl.drawing.image import Image

class ExportPOMaterialController(http.Controller):

    @http.route('/export/order_po/<int:order_id>', type='http', auth='user')
    def export_po_material_excel(self, order_id, **kwargs):
        try:
            order = request.env['warehouse.order'].sudo().browse(order_id)
            if not order or not order.exists():
                return request.make_response(
                    "Không tìm thấy chương trình hoặc chương trình không tồn tại.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            # Đường dẫn đến file template
            module_path = os.path.dirname(os.path.abspath(__file__))
            template_path = os.path.join(module_path, '../static/src/xlsx/export_order_management/export_supplier_material_order.xlsx')
            if not os.path.exists(template_path):
                return request.make_response(
                    "Không tìm thấy file mẫu export_supplier_material_order.xlsx.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            wb = load_workbook(template_path)
            ws = wb.active

            # Thêm mã chương trình vào A7 (hàng 7, cột 1)
            ws.merge_cells('C7:E7')
            ws['C7'] = order.name or ''
            ws['C7'].alignment = ws['C7'].alignment.copy(horizontal='center', vertical='center')

            # Ghi tên các PO từ C8 trở đi
            if not order.po_ids:
                return request.make_response(
                    "Chương trình chưa có PO vật tư nào.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )
            for i, po in enumerate(order.po_ids, start=3):  # C là cột 3
                col_letter = get_column_letter(i)
                ws[f'{col_letter}8'] = po.name or ''
                ws[f'{col_letter}8'].font = Font(size=8)
                ws[f'{col_letter}9'] = po.supplier_id.name_supplier or ''
                ws[f'{col_letter}9'].font = Font(size=8)

            # --- Tự động tạo header ---
            ordered_headers = [
                'Order Supplier','Mtr#', 'Type', 'Mtr.Code', 'Material name', 'Dimension','Color#',
                'Color name', 'Color set', 'Rate','Supplier#', 'Supplier', 'Cons. Qty', 'Est. Qty',
                'Price', 'Cif_price', 'Fob_price', 'Exwork_price', 'Total'
            ]
            
            header_font = Font(bold=True, size=8)
            header_map = {}
            start_row = 11
            
            # Xóa header cũ và ghi header mới
            for cell in ws[start_row]:
                cell.value = None
            for col_idx, header_title in enumerate(ordered_headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=header_title)
                cell.font = header_font
                header_map[header_title] = col_idx

            # Dòng bắt đầu ghi dữ liệu
            current_row = start_row + 1
            data_font = Font(size=8)
            has_line = False

            for po in order.po_ids:
                for line in po.line_ids:
                    has_line = True
                    row_data = {
                        'Order Supplier': po.name or '',
                        'Mtr#': line.name or '',
                        'Type': line.mtr_type.name if line.mtr_type else '',
                        'Mtr.Code': line.mtr_code or '',
                        'Material name': line.mtr_name or '',
                        'Dimension': line.dimension or '',
                        'Color#': line.color_item or '',
                        'Color name': line.color_name or '',
                        'Color set': line.color_set or '',
                        'Rate': line.rate or '',
                        'Supplier#': line.supplier.supplier_index or '',
                        'Supplier': line.supplier.name_supplier if line.supplier else '',
                        'Cons. Qty': line.cons_qty or 0,
                        'Est. Qty': line.est_qty or 0,
                        'Price': line.price or 0,
                        'Cif_price': line.cif_price or 0,
                        'Fob_price': line.fob_price or 0,
                        'Exwork_price': line.exwork_price or 0,
                        'Total': line.total or 0,
                    }

                    # Ghi dòng vào excel
                    for header, col_idx in header_map.items():
                        cell = ws.cell(row=current_row, column=col_idx, value=row_data.get(header, ''))
                        cell.font = data_font
                    
                    current_row += 1


            if not has_line:
                return request.make_response(
                    "Không có dòng vật tư nào trong các PO của chương trình này.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            # Lưu file vào bộ nhớ
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_data = output.read()
            wb.close()

            filename = f'PO_Material_List_{order.name or order.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
            return request.make_response(file_data, headers)
        except Exception as e:
            return request.make_response(
                f"Lỗi khi xuất file Excel: {str(e)}",
                [('Content-Type', 'text/plain; charset=utf-8')]
            )