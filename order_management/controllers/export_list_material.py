from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl import load_workbook
import os

class WarehouseOrderExportController(http.Controller):

    @http.route('/export/order/<int:order_id>', type='http', auth='user')
    def export_order_material_excel(self, order_id, **kwargs):
        order = request.env['warehouse.order'].sudo().browse(order_id)
        lines = order.aggregated_material_ids  # đúng tên field compute vật tư tổng hợp

        # Đường dẫn file template
        module_path = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(module_path, '../static/src/xlsx/form_export_material_order.xlsx')

        if not os.path.exists(template_path):
            return request.not_found()

        workbook = load_workbook(template_path)
        sheet = workbook.active

        # Ghi tên đơn hàng vào ô K4 (hàng 4, cột 11)
        sheet.cell(row=4, column=11, value=order.name or '')

        # Ghi danh sách mã hàng (product_code) vào hàng 9, từ cột P (16) trở đi
        for idx, code in enumerate(order.product_code_ids):
            sheet.cell(row=9, column=16 + idx, value=code.name)

        # Ghi danh sách PO vào hàng 10, từ cột P trở đi
        for idx, po in enumerate(order.po_ids):
            sheet.cell(row=10, column=16 + idx, value=po.name)
            
        # Ghi danh sách nhà cung cấp theo hàng ngang từ hàng 11, cột P
        for idx, po in enumerate(order.po_ids):
            supplier_name = po.supplier_id.name if po.supplier_id else ''
            sheet.cell(row=11, column=16 + idx, value=supplier_name)
            
        # Ghi dữ liệu vật tư bắt đầu từ dòng 14
        start_row = 14
        for idx, line in enumerate(lines):
            row = start_row + idx
            sheet.cell(row=row, column=1, value=line.mtr_type.name if line.mtr_type else '')
            sheet.cell(row=row, column=2, value=line.mtr_name or '')
            sheet.cell(row=row, column=3, value=line.mtr_code or '')
            sheet.cell(row=row, column=4, value=line.mtr_no or '')
            sheet.cell(row=row, column=5, value=line.dimension or '')
            sheet.cell(row=row, column=6, value=line.color_item or '')
            sheet.cell(row=row, column=7, value=line.color_name or '')
            sheet.cell(row=row, column=8, value=line.color_set or '')
            sheet.cell(row=row, column=9, value=line.color_code or '')
            sheet.cell(row=row, column=10, value=line.rate or '')
            sheet.cell(row=row, column=11, value=line.price or 0)
            sheet.cell(row=row, column=12, value=line.supplier.name if line.supplier else '')
            sheet.cell(row=row, column=13, value=line.est_qty or 0)
            sheet.cell(row=row, column=14, value=line.act_qty or 0)

        # Trả file về trình duyệt
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f'order_material_export_{datetime.now().strftime("%H.%M.%S.%d.%m")}.xlsx'
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename={filename}')
            ]
        )
