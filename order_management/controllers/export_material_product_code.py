from odoo import http
from odoo.http import request
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from odoo.modules.module import get_module_path
from openpyxl import load_workbook
import os
import io
from datetime import datetime

class MaterialProductExportController(http.Controller):
    
    @http.route('/export/material_product/<int:product_code_id>', type='http', auth='user')
    def export_material_product_excel(self, product_code_id, **kwargs):
        try:
            product_code = request.env['product.code'].sudo().browse(product_code_id)
            if not product_code.exists():
                return request.make_response("Style không tồn tại.", [('Content-Type', 'text/plain; charset=utf-8')])

            # 1. Sử dụng template của chức năng import để đảm bảo nhất quán
            module_path = get_module_path('order_management')
            template_path = os.path.join(module_path, 'static/src/xlsx/export_material_product/export_material_product_code.xlsx')
            if not os.path.exists(template_path):
                return request.make_response("Không tìm thấy file mẫu.", [('Content-Type', 'text/plain; charset=utf-8')])

            wb = load_workbook(template_path)
            ws = wb.active
            
            # Định nghĩa các kiểu Font
            header_font = Font(bold=True, size=8)
            data_font = Font(size=8)
            
            # 2. Ghi thông tin chung của Style
            ws['C4'] = product_code.warehouse_order_id.name  # Progarm
            ws['C5'] = product_code.name or ''               # Style#
            ws['C6'] = product_code.ean_no or ''             # Buyer Style# :
            
            ws['E5'] = product_code.description or ''        #  Style name
            ws['E6'] = product_code.customer_id.name_customer or ''     # Customer#

            ws['G5'] = product_code.total_order_qty or ''    # Total Order Qty
            ws['G6'] = product_code.total_test_qty or ''      # Total Test Qty

            info_cells = ['C4', 'C5', 'C6', 'E5', 'E6', 'G5', 'G6']
            for cell_ref in info_cells:
                ws[cell_ref].font = data_font

            # 3. Tự động tạo header mới theo thứ tự yêu cầu
            style_variants = product_code.color_size_ids
            style_sizes = style_variants.mapped('size_id').sorted(key=lambda s: s.name)
            '''
            # >> Kiểm tra xem có dữ liệu cho các cột giá tùy chọn không
            all_related_prog_custs = request.env['program.customer'].sudo().search([
                ('color_size_ids', 'in', style_variants.ids)
            ])
            all_price_lines = all_related_prog_custs.mapped('price_line_ids')
            has_cif_price = any(pl.cif_price for pl in all_price_lines)
            has_fob_price = any(pl.fob_price for pl in all_price_lines)
            has_exwork_price = any(pl.exwork_price for pl in all_price_lines)
            # <<
            '''
            # Xây dựng danh sách header theo thứ tự
            ordered_headers = ['Color.style#', 'Color.style Name']
            ordered_headers.extend([f'Size_{s.name}' for s in style_sizes])
            ordered_headers.extend(['Mtr#','Type', 'Mtr.Code', 'Material name', 'Dimension'])
            ordered_headers.extend(['Color#','Color name', 'Color set', 'Rate','Supplier#' ,'Supplier'])
            ordered_headers.extend(['Price', 'Cif_price', 'Fob_price', 'Exwork_price'])
            ordered_headers.extend([f'Cons_{s.name}' for s in style_sizes])

            # Ghi header mới vào dòng 7, áp dụng font và tạo header_map
            header_map = {}
            for cell in ws[7]: # Xóa header cũ
                cell.value = None
            for col_idx, header_title in enumerate(ordered_headers, 1):
                cell = ws.cell(row=7, column=col_idx, value=header_title)
                cell.font = header_font
                header_map[header_title] = col_idx

            # 4. Thu thập và tái cấu trúc dữ liệu
            style_colors = style_variants.mapped('color_id').sorted(key=lambda c: c.name)
            
            # Lấy toàn bộ program_customer liên quan đến style này để tối ưu
            all_related_prog_custs = request.env['program.customer'].sudo().search([
                ('color_size_ids', 'in', style_variants.ids)
            ])

            all_rows_data = []
            for color in style_colors:
                # Lấy các variant (product.color.size) cho màu hiện tại trong style này
                variants_for_color = style_variants.filtered(lambda v: v.color_id == color)
                
                # Lọc ra các program_customer áp dụng cho màu này
                relevant_prog_custs = all_related_prog_custs.filtered(
                    lambda pc: any(v in pc.color_size_ids for v in variants_for_color)
                )

                for prog_cust in relevant_prog_custs:
                    row_data = {
                        'Color.style#': color.color_code,
                        'Color.style Name': color.name,
                        'Mtr#': prog_cust.name or '',
                        'Type': prog_cust.mtr_type.name if prog_cust.mtr_type else '',
                        'Material name': prog_cust.mtr_name,
                        'Mtr.Code': prog_cust.mtr_code,
                        'Dimension': prog_cust.dimension,
                        'Color#': prog_cust.material_color_id.name if prog_cust.material_color_id else '',
                        'Color name': prog_cust.material_color_id.color_name if prog_cust.material_color_id else '',
                        'Color set': prog_cust.material_color_id.color_set_id.name if prog_cust.material_color_id.color_set_id else '',
                        'Rate': prog_cust.rate,
                        'Supplier#': prog_cust.supplier.supplier_index if prog_cust.supplier else '',
                        'Supplier': prog_cust.supplier.name_supplier if prog_cust.supplier else '',
                        'Price': prog_cust.price,
                        'Cif_price': prog_cust.cif_price,
                        'Fob_price': prog_cust.fob_price,
                        'Exwork_price': prog_cust.exwork_price,                        
                    }

                    # Lấy dữ liệu chi tiết theo từng size
                    for size in style_sizes:
                        # Tìm variant cụ thể cho color và size này
                        variant = style_variants.filtered(lambda v: v.color_id == color and v.size_id == size)
                        
                        # Kiểm tra xem vật tư (prog_cust) này có được áp dụng cho variant (size) này không
                        if not (variant and variant in prog_cust.color_size_ids):
                            continue

                        size_name = size.name
                        row_data[f'Size_{size_name}'] = 'x'
                        
                        norm_line = prog_cust.norm_line_ids.filtered(lambda n: n.color_size_id == variant)
                        row_data[f'Cons_{size_name}'] = norm_line.consumption if norm_line else 0.0

                    all_rows_data.append(row_data)
                    
            # 5. Ghi dữ liệu vào sheet
            current_row = 8
            for data in all_rows_data:
                for header, value in data.items():
                    col_idx = header_map.get(header)
                    if col_idx:
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.value = value
                        cell.font = data_font
                current_row += 1

            # 6. Lưu và trả về file
            output = io.BytesIO()
            wb.save(output)
            file_data = output.getvalue()

            filename = f'Export_{product_code.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.xlsx'
            headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
            return request.make_response(file_data, headers)

        except Exception as e:
            return request.make_response(f"Lỗi khi xuất file Excel: {str(e)}", [('Content-Type', 'text/plain; charset=utf-8')])