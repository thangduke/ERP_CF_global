from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from odoo.modules.module import get_module_path
from openpyxl import load_workbook
import os

class ProductCodeExportController(http.Controller):
    _description = 'Export danh sách vật tư trong giao diện style(color,size)'

    @http.route('/export/product/<int:product_color_size_id>', type='http', auth='user')
    def export_material_product_color_size_excel(self, product_color_size_id, **kwargs):
        try:
            product_color_size = request.env['product.color.size'].sudo().browse(product_color_size_id)
            if not product_color_size.exists():
                return request.make_response(
                    "Không tìm thấy Style-Color-Size hoặc bản ghi không tồn tại.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )
            
            product_code = product_color_size.product_code_id
            if not product_code:
                return request.make_response(
                    "Bản ghi Style-Color-Size không được liên kết với một Style (Product Code).",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            module_path = get_module_path('order_management')
            template_path = os.path.join(module_path, 'static/src/xlsx/export_product_color_size/export_material_product_color_size.xlsx')
            if not os.path.exists(template_path):
                return request.make_response(
                    "Không tìm thấy file mẫu export_material_product_color_size.xlsx.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )
            
            wb = load_workbook(template_path)
            ws = wb.active

            header_font = Font(bold=True, size=8)
            data_font = Font(size=8)

            # Ghi thông tin chung
            ws['C7'] = product_code.warehouse_order_id.name or ''
            ws['C8'] = product_code.name or ''
            ws['C9'] = product_code.customer_id.name_customer or ''
            ws['F7'] = product_color_size.color or ''
            ws['F8'] = product_color_size.color_id.name or ''
            ws['F9'] = product_color_size.size_id.name or ''
            ws['H7'] = product_color_size.order_qty or ''
            ws['H8'] = product_color_size.test_qty or ''
            
            info_cells = ['C7', 'C8', 'C9', 'F7', 'F8', 'F9', 'H7', 'H8']
            for cell_ref in info_cells:
                ws[cell_ref].font = data_font

            # Chuẩn bị header động
            style_variants = product_code.color_size_ids
            style_sizes = style_variants.mapped('size_id').sorted(key=lambda s: s.name)

            ordered_headers = [
                'Position','Mtr#', 'Type', 'Mtr.Code', 'Material name', 'Dimension', 'Color#',
                'Color name', 'Color set', 'Rate', 'Supplier#', 'Supplier',
                'Consumption', 'Price', 'Cif_price', 'Fob_price', 'Exwork_price', 'Total'
            ]

            # Ghi header
            header_map = {}
            for cell in ws[11]: # Xóa header cũ trên dòng 11
                cell.value = None
            for col_idx, header_title in enumerate(ordered_headers, 1):
                cell = ws.cell(row=11, column=col_idx, value=header_title)
                cell.font = header_font
                header_map[header_title] = col_idx

            # Thu thập và xử lý dữ liệu
            current_row = 12
            
            materials_with_context = product_color_size.material_ids.with_context(
                active_model='product.color.size',
                active_id=product_color_size.id
            )

            if materials_with_context:
                for material in materials_with_context:
                    row_data = {
                        'Position': material.position or '',
                        'Mtr#': material.name or '',
                        'Type': material.mtr_type.name if material.mtr_type else '',
                        'Mtr.Code': material.mtr_code or '',
                        'Material name': material.mtr_name or '',
                        'Dimension': material.dimension or '',
                        'Color#': material.color_item or '',
                        'Color name': material.color_name or '',
                        'Color set': material.color_set or '',
                        'Rate': material.rate or '',
                        'Supplier#': material.supplier.supplier_index if material.supplier else '',
                        'Supplier': material.supplier.name_supplier if material.supplier else '',
                        'Consumption': material.consumption or 0,
                        'Price': material.price or 0,
                        'Cif_price': material.cif_price or 0,
                        'Fob_price': material.fob_price or 0,
                        'Exwork_price': material.exwork_price or 0,
                        'Total': material.total or 0,
                    }

                    # Ghi dòng vào excel
                    for header, col_idx in header_map.items():
                        cell = ws.cell(row=current_row, column=col_idx, value=row_data.get(header, ''))
                        cell.font = data_font
                    
                    current_row += 1

            # Lưu và trả về file
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_data = output.read()
            wb.close()

            filename = f'Product_{product_code.display_name or product_code.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
            return request.make_response(file_data, headers)
        except Exception as e:
            return request.make_response(
                f"Lỗi khi xuất file Excel: {str(e)}",
                [('Content-Type', 'text/plain; charset=utf-8')]
            )
