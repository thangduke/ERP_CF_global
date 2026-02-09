from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from openpyxl.drawing.image import Image


class ManufacturingOrderExportController(http.Controller):

    @http.route('/export/manufacturing_order/<int:order_id>', type='http', auth='user')
    def export_manufacturing_order_excel(self, order_id, **kwargs):
        order = request.env['warehouse.order'].sudo().browse(order_id)
        if not order:
            return request.not_found()

        # Đường dẫn đến file template gốc
        module_path = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(module_path, '../static/src/xlsx/export_order_management/export_manufacturing_order.xlsx')


        if not os.path.exists(template_path):
            return request.not_found()

        # Load file mẫu
        workbook = load_workbook(template_path)
        template_sheet = workbook.active

        for product in order.product_code_ids:
            sheet = workbook.copy_worksheet(template_sheet)
            sheet.title = (product.description or 'MO')[:31]

            logo_path = os.path.join(module_path, '../static/src/xlsx/logo.png')
            img = Image(logo_path)
            sheet.add_image(img, 'A1')  # Vị trí A1, tuỳ chỉnh theo template
            # Ghi thông tin chung cho lệnh sản xuất
            sheet['B6'] = order.name or ''  # Tên chương trình
            sheet['B7'] = product.name or ''  # Style
            sheet['B8'] = order.customer_id.name_customer or ''  # Khách hàng
            sheet['B9'] = order.customer_po_index or ''  # Mã PO khách hàng
            sheet['B11'] = product.description or ''  # Mô tả Style

            #sheet['M6'] = order.factory_id.name or ''  # Nhà máy sản xuất

            sheet['M8'] = product.description or ''  # Mô tả Style (nếu cần)
            sheet['M9'] = (product.total_order_qty or 0) + (product.total_test_qty or 0)  # Tổng số lượng
            sheet['M10'] = order.order_date.strftime('%d/%m/%Y') if order.order_date else ''
            sheet['M11'] = order.ship_date.strftime('%d/%m/%Y') if order.ship_date else ''
            # Thêm tên người tạo vào O:51
            sheet['O51'] = order.employee_id.name or ''

            # Thêm tên người duyệt vào J,K,L:51 (gộp & căn giữa)
            sheet.merge_cells('J51:L51')
            sheet['J51'] = ', '.join(order.manager_ids.mapped('name')) if order.manager_ids else ''
            sheet['J51'].alignment = sheet['J51'].alignment.copy(horizontal='center', vertical='center')

            # Thêm phòng ban khác duyệt vào D,E,F,G:51 (gộp & căn giữa)
            sheet.merge_cells('D51:G51')
            sheet['D51'] = order.department_approval_id.name or ''
            sheet['D51'].alignment = sheet['D51'].alignment.copy(horizontal='center', vertical='center')
            # Lấy danh sách size theo thứ tự
            size_objs = product.color_size_ids.mapped('size_id')
            size_objs = sorted(size_objs, key=lambda s: s.sequence if hasattr(s, 'sequence') else 999, reverse=True)
            size_names = [s.name for s in size_objs]

            # Lấy danh sách màu (unique theo color_id)
            color_objs = product.color_size_ids.mapped('color_id')
            color_objs = list({c.id: c for c in color_objs}.values())

            row_header = 15
            row_data_start = 16
            default_size_count = 10
            actual_size_count = len(size_objs)

            if actual_size_count > default_size_count:
                total_col_idx = 3 + default_size_count
                extra_cols = actual_size_count - default_size_count
                sheet.insert_cols(total_col_idx, amount=extra_cols)

            # Ghi lại tiêu đề size
            for idx, size_name in enumerate(size_names):
                sheet.cell(row=row_header, column=3 + idx, value=size_name)

            # Ghi dữ liệu và công thức tổng cho từng dòng màu
            for row_idx, color in enumerate(color_objs):
                row = row_data_start + row_idx
                sheet.cell(row=row, column=1, value=color.name or '')
                sheet.cell(row=row, column=2, value=color.color_code or '')

                for col_idx, size in enumerate(size_objs):
                    qty = 0
                    color_size = product.color_size_ids.filtered(lambda cs: cs.color_id.id == color.id and cs.size_id.id == size.id)
                    if color_size:
                        qty = color_size[0].order_qty or 0
                    sheet.cell(row=row, column=3 + col_idx, value=qty)

                # Công thức tổng động
                first_size_col = get_column_letter(3)
                last_size_col = get_column_letter(2 + actual_size_count)
                if actual_size_count <= default_size_count:
                    total_col_idx = 3 + default_size_count  # Luôn là cột M
                else:
                    total_col_idx = 3 + actual_size_count   # Cột tổng dịch sang phải
                total_col = get_column_letter(total_col_idx)
                sheet.cell(row=row, column=total_col_idx, value=f"=SUM({first_size_col}{row}:{last_size_col}{row})")
                
            # --- GHI BẢNG TEST ---    
            row_header_test = 24  
            row_data_start_test = 25

            # Tiêu đề bảng test
            sheet.cell(row=row_header_test - 2, column=1, value="MẪU TOP+LƯU+TEST")
            for idx, size_name in enumerate(size_names):
                sheet.cell(row=row_header_test, column=3 + idx, value=size_name)
            sheet.cell(row=row_header_test, column=1, value="COLOR NAME")
            sheet.cell(row=row_header_test, column=2, value="COLOR CODE")

            # Ghi dữ liệu test_qty cho từng dòng màu
            for row_idx, color in enumerate(color_objs):
                row = row_data_start_test + row_idx
                sheet.cell(row=row, column=1, value=color.name or '')
                sheet.cell(row=row, column=2, value=color.color_code or '')

                for col_idx, size in enumerate(size_objs):
                    qty = 0
                    color_size = product.color_size_ids.filtered(lambda cs: cs.color_id.id == color.id and cs.size_id.id == size.id)
                    if color_size:
                        qty = color_size[0].test_qty or 0  # Số lượng test
                    sheet.cell(row=row, column=3 + col_idx, value=qty)

                # Công thức tổng động cho bảng test
                first_size_col = get_column_letter(3)
                last_size_col = get_column_letter(2 + actual_size_count)
                if actual_size_count <= default_size_count:
                    total_col_idx = 3 + default_size_count
                else:
                    total_col_idx = 3 + actual_size_count
                total_col = get_column_letter(total_col_idx)
                sheet.cell(row=row, column=total_col_idx, value=f"=SUM({first_size_col}{row}:{last_size_col}{row})")
                
                
        # Xóa sheet gốc template nếu không cần giữ lại
        if template_sheet.title in workbook.sheetnames:
            workbook.remove(template_sheet)

        # Xuất file Excel
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f'ManufacturingOrders_{order.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename={filename}')
            ]
        )