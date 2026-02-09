from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment
import os
from openpyxl.drawing.image import Image

class MaterialInvoiceExportController(http.Controller):
    
    @http.route('/export/material_invoice/<int:invoice_id>', type='http', auth='user')
    def export_material_excel(self, invoice_id, **kwargs):
        try:
            invoice = request.env['material.invoice'].sudo().browse(invoice_id)
            if not invoice.exists():
                return request.make_response("Không tìm thấy Mã lọc vật tư theo NCC.",
                                             [('Content-Type', 'text/plain; charset=utf-8')])

            # Đường dẫn đến file template
            module_path = os.path.dirname(os.path.abspath(__file__))
            template_path = os.path.join(module_path, '../static/src/xlsx/export_material_PO/export_material_invoice.xlsx')
            if not os.path.exists(template_path):
                return request.make_response(
                    "Không tìm thấy file mẫu export_material_invoice.xlsx",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            wb = load_workbook(template_path)
            ws = wb.active

            # Thêm mã chương trình vào A7 (hàng 7, cột 1)
            ws.merge_cells('C6:E6')
            ws['C6'] = invoice.order_id.name or ''
            ws['C6'].alignment = ws['C6'].alignment.copy(horizontal='center', vertical='center')

            # Ghi tên các PO từ C8 trở đi
            ws['C7'] = invoice.po_id.display_name or ''
            ws['C7'].alignment = ws['C7'].alignment.copy(horizontal='center', vertical='center')
            
            ws['C8'] = invoice.supplier.name_supplier or ''
            ws['C8'].alignment = ws['C8'].alignment.copy(horizontal='center', vertical='center')
        
            
            ws['C9'] = invoice.name or ''
            ws['C9'].alignment = ws['C9'].alignment.copy(horizontal='center', vertical='center')
            # Dòng bắt đầu ghi dữ liệu (ví dụ: dòng 11)
            # Dòng bắt đầu ghi dữ liệu
            header_row = 10
            start_row = 11

            # --- Định nghĩa và ghi headers ---
            ordered_headers = [
                'Mtr#', 'Type', 'Mtr.Code', 'Material name', 'Dimension', 'Color name', 'Color set', 'Rate', 'Supplier',
                'Est.Qty', 'PO.Qty', 'Inv.Qty', 'Price', 'Cif_price', 'Fob_price', 'Exwork_price', 'Total'
            ]
            header_font = Font(bold=True)
            for col_num, header in enumerate(ordered_headers, 1):
                cell = ws.cell(row=header_row, column=col_num, value=header)
                cell.font = header_font

            # --- Ghi dữ liệu ---
            data_font = Font(size=8)
            idx = start_row
            has_line = False
            for line in invoice.invoice_line_ids:
                has_line = True
                
                data_row = [
                    line.name or '',
                    line.mtr_type.name if line.mtr_type else '',
                    line.mtr_code or '',
                    line.mtr_name or '',
                    line.dimension or '',
                    line.color_name or '',
                    line.color_set or '',
                    line.rate or '',
                    line.supplier.name_supplier or '',
                    line.est_qty or 0,
                    line.act_qty or 0,
                    line.inv_qty or 0,
                    line.price or 0,
                    line.cif_price or 0,
                    line.fob_price or 0,
                    line.exwork_price or 0,
                    line.total or 0
                ]

                for col_num, value in enumerate(data_row, 1):
                    cell = ws.cell(row=idx, column=col_num, value=value)
                    cell.font = data_font

                idx += 1


            if not has_line:
                return request.make_response(
                    "Không có dòng vật tư nào trong các PO của chương trình này.",
                    [('Content-Type', 'text/plain; charset=utf-8')]
                )

            # Lưu file vào bộ nhớ
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_data = output.read()
            wb.close()

            filename = f'PO_Material_List_{invoice.supplier.name_supplier or invoice.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
            return request.make_response(file_data, headers)
        except Exception as e:
            return request.make_response(
                f"Lỗi khi xuất file Excel: {str(e)}",
                [('Content-Type', 'text/plain; charset=utf-8')]
            )

