from odoo import http
from odoo.http import request
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import shutil

class WarehousePOExportController(http.Controller):

    @http.route('/export/order_po/<int:order_id>', type='http', auth='user')
    def export_po_material_excel(self, order_id, **kwargs):
        order = request.env['warehouse.order'].sudo().browse(order_id)
        if not order:
            return request.not_found()

        # Đường dẫn đến file template gốc
        module_path = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(module_path, '../static/src/xlsx/form_export_material_po.xlsx')

        if not os.path.exists(template_path):
            return request.not_found()

        # Load file mẫu
        workbook = load_workbook(template_path)
        template_sheet = workbook.active

        # Tên đơn hàng
        order_name = order.name or 'Đơn hàng'

        for po in order.po_ids:
            # Tạo sheet mới cho mỗi PO
            po_sheet = workbook.copy_worksheet(template_sheet)
            po_sheet.title = po.name[:31]  # Excel giới hạn tên sheet tối đa 31 ký tự

            # Ghi thông tin chung
            po_sheet.cell(row=7, column=3, value=po.name or '')
            po_sheet.cell(row=8, column=3, value=po.supplier_id.name or '')
            po_sheet.cell(row=9, column=3, value=po.date_order.strftime('%d/%m/%Y') if po.date_order else '')
            po_sheet.cell(row=11, column=3, value=order_name)

            # Ghi dữ liệu chi tiết vật tư
            row_start = 14
            for idx, line in enumerate(po.line_ids):
                row = row_start + idx
                po_sheet.cell(row=row, column=2, value=line.mtr_code or '')
                po_sheet.cell(row=row, column=3, value=line.mtr_name or '')
                po_sheet.cell(row=row, column=4, value=line.dimension or '')
                po_sheet.cell(row=row, column=5, value=line.color_name or '')
                po_sheet.cell(row=row, column=6, value=line.color_set or '')
                po_sheet.cell(row=row, column=7, value=line.color_code or '')
                po_sheet.cell(row=row, column=8, value=line.est_qty or 0.0)
                po_sheet.cell(row=row, column=9, value=line.rate or '')
                po_sheet.cell(row=row, column=10, value=line.price or 0.0)
                po_sheet.cell(row=row, column=11, value=line.total or 0.0)

        # Xóa sheet gốc template nếu không cần giữ lại
        if 'Sheet' in workbook.sheetnames:
            std_sheet = workbook['Sheet']
            workbook.remove(std_sheet)
        elif template_sheet.title in workbook.sheetnames:
            workbook.remove(template_sheet)

        # Xuất file Excel
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f'POs_{order_name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename={filename}')
            ]
        )
