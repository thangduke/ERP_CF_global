import io
from odoo import http
from odoo.http import request, content_disposition
import openpyxl
from openpyxl.styles import Font, Alignment

class EmployeeExportController(http.Controller):

    @http.route('/export/employee_list', type='http', auth='user')
    def export_employee_list_excel(self, **kwargs):
        try:
            employees = request.env['employee.base'].sudo().search([])
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh sách nhân sự"

            # --- Định nghĩa Style ---
            header_font = Font(name='Times New Roman', bold=True, size=12)
            cell_font = Font(name='Times New Roman', size=11)
            center_alignment = Alignment(horizontal='center', vertical='center')

            # --- Tạo Header (khớp với file import) ---
            headers = [
                "Họ và tên", "Tài khoản", "Mã nhân sự", "Bộ phận", "Vị trí công việc", 
                "Vị trí công tác", "Loại nhân sự", "Ngày bắt đầu", "Ngày chính thức", 
                "Giới tính", "Ngày sinh", "Số điện thoại"
            ]
            ws.append(headers)

            # --- Style cho Header ---
            for col_num, header_title in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.alignment = center_alignment
                ws.column_dimensions[cell.column_letter].autosize = True

            # --- Ghi dữ liệu ---
            gender_map = {'male': 'Nam', 'female': 'Nữ', 'other': 'Khác'}
            for emp in employees:
                row_data = [
                    emp.name or '',
                    emp.user_id.login or '',
                    emp.employee_index or '',
                    emp.department_id.name or '',
                    emp.position_id.name or '',
                    emp.position_type.name or '',
                    emp.employee_type_2.name or '',
                    emp.start_date.strftime('%d/%m/%Y') if emp.start_date else '',
                    emp.official_date.strftime('%d/%m/%Y') if emp.official_date else '',
                    gender_map.get(emp.gender, ''),
                    emp.birthday.strftime('%d/%m/%Y') if emp.birthday else '',
                    emp.mobile_phone or '',
                ]
                ws.append(row_data)
            
            # --- Style cho các ô dữ liệu ---
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.font = cell_font

            # --- Tạo file và trả về ---
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            file_name = "Danh_sach_nhan_su.xlsx"
            
            return request.make_response(
                output.read(),
                [
                    ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                    ('Content-Disposition', content_disposition(file_name))
                ]
            )

        except Exception as e:
            return request.make_response(
                f"Đã có lỗi xảy ra: {e}",
                [('Content-Type', 'text/plain; charset=utf-8')]
            )